"""Automação de planilha Excel com openpyxl.

Gera uma planilha de vendas com dados de exemplo, aplica fórmulas,
formatação e cria um resumo automático.
"""

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PASTA = Path(__file__).parent
ARQUIVO_SAIDA = PASTA / "relatorio_vendas.xlsx"

PRODUTOS = ["Notebook", "Mouse", "Teclado", "Monitor", "Headset", "Webcam"]
VENDEDORES = ["Ana", "Bruno", "Carla", "Diego", "Elisa"]


def gerar_dados(quantidade: int = 30) -> list[dict]:
    dados = []
    for _ in range(quantidade):
        produto = random.choice(PRODUTOS)
        preco = round(random.uniform(50, 3500), 2)
        quantidade_vendida = random.randint(1, 10)
        dados.append({
            "vendedor": random.choice(VENDEDORES),
            "produto": produto,
            "preco_unitario": preco,
            "quantidade": quantidade_vendida,
        })
    return dados


def criar_planilha(dados: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    cabecalho = ["Vendedor", "Produto", "Preço Unitário", "Quantidade", "Total"]
    ws.append(cabecalho)

    fonte_titulo = Font(bold=True, color="FFFFFF")
    fundo_titulo = PatternFill("solid", fgColor="305496")
    for col, _ in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = fonte_titulo
        cell.fill = fundo_titulo
        cell.alignment = Alignment(horizontal="center")

    for i, venda in enumerate(dados, start=2):
        ws.cell(row=i, column=1, value=venda["vendedor"])
        ws.cell(row=i, column=2, value=venda["produto"])
        ws.cell(row=i, column=3, value=venda["preco_unitario"]).number_format = "R$ #,##0.00"
        ws.cell(row=i, column=4, value=venda["quantidade"])
        ws.cell(row=i, column=5, value=f"=C{i}*D{i}").number_format = "R$ #,##0.00"

    ultima_linha = len(dados) + 1
    linha_total = ultima_linha + 2
    ws.cell(row=linha_total, column=4, value="Total Geral:").font = Font(bold=True)
    ws.cell(row=linha_total, column=5, value=f"=SUM(E2:E{ultima_linha})").number_format = "R$ #,##0.00"
    ws.cell(row=linha_total, column=5).font = Font(bold=True)

    for col in range(1, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(ARQUIVO_SAIDA)
    print(f"Planilha gerada: {ARQUIVO_SAIDA}")
    print(f"Total de registros: {len(dados)}")


if __name__ == "__main__":
    criar_planilha(gerar_dados())
